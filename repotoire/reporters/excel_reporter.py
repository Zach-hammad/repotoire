"""Excel report generator for Repotoire analysis results.

Generates multi-sheet Excel workbooks with summary, findings, and metrics data.
Requires openpyxl: pip install openpyxl
"""

from datetime import datetime
from pathlib import Path
from typing import List, Optional

from repotoire.models import CodebaseHealth, Finding, Severity
from repotoire.logging_config import get_logger
from repotoire.reporters.base_reporter import BaseReporter

logger = get_logger(__name__)


class ExcelReporter(BaseReporter):
    """Generate Excel reports from analysis results.

    Inherits common functionality from BaseReporter including code snippet
    extraction and language detection.
    """

    def __init__(self, repo_path: Path | str | None = None):
        """Initialize Excel reporter.

        Args:
            repo_path: Path to repository (for reference in report)
        """
        super().__init__(repo_path=repo_path, include_snippets=False)

    def generate(self, health: CodebaseHealth, output_path: Path) -> None:
        """Generate Excel report from health data.

        Args:
            health: CodebaseHealth instance with analysis results
            output_path: Path to output Excel file

        Raises:
            ImportError: If openpyxl is not installed
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference, PieChart
        except ImportError:
            raise ImportError(
                "Excel report generation requires openpyxl. "
                "Install with: pip install openpyxl"
            )

        wb = Workbook()

        # Create sheets
        self._create_summary_sheet(wb, health, Font, PatternFill, Alignment, Border, Side)
        self._create_findings_sheet(wb, health, Font, PatternFill, Alignment, Border, Side)
        self._create_metrics_sheet(wb, health, Font, PatternFill, Alignment, Border, Side)
        self._create_detectors_sheet(wb, health, Font, PatternFill, Alignment)

        # Remove default sheet if we created others
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"Excel report generated: {output_path}")

    def _create_summary_sheet(self, wb, health, Font, PatternFill, Alignment, Border, Side):
        """Create summary sheet with overview."""
        ws = wb.create_sheet("Summary", 0)

        # Styles
        header_font = Font(bold=True, size=14)
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF")
        grade_fonts = {
            "A": Font(bold=True, size=24, color="00B050"),
            "B": Font(bold=True, size=24, color="92D050"),
            "C": Font(bold=True, size=24, color="FFEB00"),
            "D": Font(bold=True, size=24, color="FF6600"),
            "F": Font(bold=True, size=24, color="FF0000"),
        }

        # Title
        ws["A1"] = "Repotoire Code Health Report"
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:D1")

        # Generation info
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(italic=True, color="808080")

        # Grade display
        ws["A4"] = "Grade"
        ws["A4"].font = header_font
        ws["B4"] = health.grade
        ws["B4"].font = grade_fonts.get(health.grade, Font(bold=True, size=24))
        ws["B4"].alignment = Alignment(horizontal="center")

        ws["A5"] = "Overall Score"
        ws["A5"].font = header_font
        ws["B5"] = f"{health.overall_score:.1f}/100"
        ws["B5"].font = Font(bold=True, size=14)

        # Category scores section
        ws["A7"] = "Category Scores"
        ws["A7"].font = Font(bold=True, size=12)
        ws["A7"].fill = title_fill
        ws["A7"].font = title_font
        ws.merge_cells("A7:C7")

        categories = [
            ("Graph Structure", health.structure_score, "40%"),
            ("Code Quality", health.quality_score, "30%"),
            ("Architecture Health", health.architecture_score, "30%"),
        ]

        ws["A8"] = "Category"
        ws["B8"] = "Score"
        ws["C8"] = "Weight"
        for col in ["A", "B", "C"]:
            ws[f"{col}8"].font = Font(bold=True)
            ws[f"{col}8"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for i, (name, score, weight) in enumerate(categories, start=9):
            ws[f"A{i}"] = name
            ws[f"B{i}"] = f"{score:.1f}"
            ws[f"C{i}"] = weight

        # Findings summary section
        ws["A13"] = "Findings Summary"
        ws["A13"].font = title_font
        ws["A13"].fill = title_fill
        ws.merge_cells("A13:C13")

        ws["A14"] = "Severity"
        ws["B14"] = "Count"
        for col in ["A", "B"]:
            ws[f"{col}14"].font = Font(bold=True)
            ws[f"{col}14"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        severity_data = [
            ("Critical", health.findings_summary.critical, "FF0000"),
            ("High", health.findings_summary.high, "FF6600"),
            ("Medium", health.findings_summary.medium, "FFEB00"),
            ("Low", health.findings_summary.low, "0066FF"),
            ("Info", health.findings_summary.info, "808080"),
        ]

        for i, (name, count, color) in enumerate(severity_data, start=15):
            ws[f"A{i}"] = name
            ws[f"B{i}"] = count
            ws[f"A{i}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if color in ["FF0000", "0066FF"]:
                ws[f"A{i}"].font = Font(color="FFFFFF")

        ws["A20"] = "Total"
        ws["B20"] = health.findings_summary.total
        ws["A20"].font = Font(bold=True)
        ws["B20"].font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _create_findings_sheet(self, wb, health, Font, PatternFill, Alignment, Border, Side):
        """Create detailed findings sheet."""
        ws = wb.create_sheet("Findings")

        # Header row
        headers = ["ID", "Severity", "Detector", "Title", "Description", "Files", "Suggested Fix", "Priority"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Severity colors
        severity_fills = {
            Severity.CRITICAL: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            Severity.HIGH: PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
            Severity.MEDIUM: PatternFill(start_color="FFEB00", end_color="FFEB00", fill_type="solid"),
            Severity.LOW: PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
            Severity.INFO: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }

        # Data rows
        for row, finding in enumerate(health.findings, start=2):
            ws.cell(row=row, column=1, value=finding.id or f"finding-{row-1}")

            sev_cell = ws.cell(row=row, column=2, value=finding.severity.value if finding.severity else "unknown")
            if finding.severity in severity_fills:
                sev_cell.fill = severity_fills[finding.severity]
                if finding.severity in [Severity.CRITICAL, Severity.HIGH]:
                    sev_cell.font = Font(color="FFFFFF")

            ws.cell(row=row, column=3, value=finding.detector or "")
            ws.cell(row=row, column=4, value=finding.title or "")
            ws.cell(row=row, column=5, value=finding.description or "")
            ws.cell(row=row, column=6, value=", ".join(finding.affected_files or [])[:500])
            ws.cell(row=row, column=7, value=finding.suggested_fix or "")
            ws.cell(row=row, column=8, value=finding.priority_score if finding.priority_score else "")

        # Adjust column widths
        widths = [15, 12, 25, 40, 60, 40, 50, 10]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Add filters
        ws.auto_filter.ref = ws.dimensions

    def _create_metrics_sheet(self, wb, health, Font, PatternFill, Alignment, Border, Side):
        """Create metrics sheet."""
        ws = wb.create_sheet("Metrics")

        m = health.metrics

        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        ws["A1"] = "Metric"
        ws["B1"] = "Value"
        ws["C1"] = "Assessment"
        for col in ["A", "B", "C"]:
            ws[f"{col}1"].fill = header_fill
            ws[f"{col}1"].font = header_font

        # Metrics data
        metrics_data = [
            ("Total Files", m.total_files, ""),
            ("Total Classes", m.total_classes, ""),
            ("Total Functions", m.total_functions, ""),
            ("", "", ""),  # Separator
            ("Modularity Score", f"{m.modularity:.3f}", "Good" if m.modularity >= 0.5 else "Fair" if m.modularity >= 0.3 else "Low"),
            ("Average Coupling", f"{m.avg_coupling:.2f}" if m.avg_coupling else "N/A", "Good" if (m.avg_coupling or 0) < 3 else "Fair" if (m.avg_coupling or 0) < 5 else "High"),
            ("Circular Dependencies", m.circular_dependencies, "None" if m.circular_dependencies == 0 else f"{m.circular_dependencies} found"),
            ("Dead Code %", f"{m.dead_code_percentage:.1%}", "Minimal" if m.dead_code_percentage < 0.05 else "Needs attention"),
            ("God Class Count", m.god_class_count, "None" if m.god_class_count == 0 else f"{m.god_class_count} found"),
            ("Duplication %", f"{m.duplication_percentage:.1%}", "Minimal" if m.duplication_percentage < 0.05 else "Needs attention"),
            ("", "", ""),  # Separator
            ("Layer Violations", m.layer_violations, "None" if m.layer_violations == 0 else f"{m.layer_violations} found"),
            ("Boundary Violations", m.boundary_violations, "None" if m.boundary_violations == 0 else f"{m.boundary_violations} found"),
            ("Abstraction Ratio", f"{m.abstraction_ratio:.2f}", "Good" if 0.3 <= m.abstraction_ratio <= 0.7 else "Needs review"),
        ]

        for row, (metric, value, assessment) in enumerate(metrics_data, start=2):
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value
            ws[f"C{row}"] = assessment

            # Highlight assessment cells
            if assessment in ["Good", "Minimal", "None"]:
                ws[f"C{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif assessment in ["Low", "High", "Needs attention", "Needs review"] or "found" in str(assessment):
                ws[f"C{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20

    def _create_detectors_sheet(self, wb, health, Font, PatternFill, Alignment):
        """Create detectors breakdown sheet."""
        ws = wb.create_sheet("By Detector")

        # Count findings by detector
        detector_counts = {}
        for finding in health.findings:
            detector = finding.detector or "Unknown"
            if detector not in detector_counts:
                detector_counts[detector] = {"total": 0, "critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
            detector_counts[detector]["total"] += 1
            if finding.severity == Severity.CRITICAL:
                detector_counts[detector]["critical"] += 1
            elif finding.severity == Severity.HIGH:
                detector_counts[detector]["high"] += 1
            elif finding.severity == Severity.MEDIUM:
                detector_counts[detector]["medium"] += 1
            elif finding.severity == Severity.LOW:
                detector_counts[detector]["low"] += 1
            else:
                detector_counts[detector]["info"] += 1

        # Header
        headers = ["Detector", "Total", "Critical", "High", "Medium", "Low", "Info"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows sorted by total findings
        sorted_detectors = sorted(detector_counts.items(), key=lambda x: x[1]["total"], reverse=True)

        for row, (detector, counts) in enumerate(sorted_detectors, start=2):
            ws.cell(row=row, column=1, value=detector.replace("Detector", ""))
            ws.cell(row=row, column=2, value=counts["total"])
            ws.cell(row=row, column=3, value=counts["critical"])
            ws.cell(row=row, column=4, value=counts["high"])
            ws.cell(row=row, column=5, value=counts["medium"])
            ws.cell(row=row, column=6, value=counts["low"])
            ws.cell(row=row, column=7, value=counts["info"])

            # Color code the severity columns
            if counts["critical"] > 0:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws.cell(row=row, column=3).font = Font(color="FFFFFF")
            if counts["high"] > 0:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        for col in ["B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 12

        # Add filters
        ws.auto_filter.ref = ws.dimensions
